import json
import os
import subprocess
import sys
from datetime import datetime, timedelta

import pandas as pd
import pygsheets
from openpyxl import load_workbook
from openpyxl.styles import Font

# 常量設置
SCRIPT_DIRECTORY = '/Users/chenyaoxuan/Desktop/myproject/MarketProject/src/scripts'
USER_INFO_PATH = '/Users/chenyaoxuan/Desktop/myproject/MarketProject/userInfo.json'
OUTPUT_BASE_DIRECTORY = '/Users/chenyaoxuan/Desktop/myproject/MarketProject/marketExcel'
GOOGLE_SHEETS_KEY = "/Users/chenyaoxuan/Desktop/myproject/MarketProject/astute-harmony-425407-p9-826d675d86f9.json"
GOOGLE_SHEETS_URL = "https://docs.google.com/spreadsheets/d/1SGNEuh-FjiBxoxH73NcDELDGbSO71vhc7DbkGuKk4wo/edit#gid=1894821213"

def execute_python_script(script_path):
    """運行指定的Python腳本並返回其輸出"""
    try:
        result = subprocess.run([sys.executable, script_path], capture_output=True, text=True, check=True)
        return result.stdout
    except subprocess.CalledProcessError as e:
        print(f"運行腳本 {script_path} 時出現錯誤: {e}")
        return None

def load_json_file(json_path):
    """從指定的JSON文件加載數據"""
    try:
        with open(json_path, 'r', encoding='utf-8') as file:
            return json.load(file)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"加載JSON文件 {json_path} 時出現錯誤: {e}")
        return None

def format_excel(file_path, font_size=20):
    """設置Excel文件中的字體大小，將「銀行卡」欄位寬度設置為50，其他欄位設置為25，行高設置為50"""
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(size=font_size)
    
    # 設置列寬
    for col in sheet.columns:
        if col[0].value == "銀行卡":
            sheet.column_dimensions[col[0].column_letter].width = 50
        else:
            sheet.column_dimensions[col[0].column_letter].width = 25
    
    # 設置行高
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 50
    
    workbook.save(file_path)

def upload_to_google_sheets(data, sheet, user_info):
    """將數據寫入Google Sheets"""
    for title, details in data.items():
        try:
            worksheet = sheet.worksheet('title', title)
        except pygsheets.WorksheetNotFound:
            # 使用 f-string 格式化添加新的工作表
            worksheet = sheet.add_worksheet(title)
        
        worksheet.clear()
        df = pd.DataFrame(details, columns=["編號姓名", "時間", "銀行卡", "提款方式", "金額", "狀態"])
        worksheet.set_dataframe(df, (1, 1))
        
        print(f"{title} 資料已成功更新")

def collect_transactions(scripts, script_directory, user_info):
    """處理所有腳本並收集交易數據"""
    all_transactions = {}

    for script in scripts:
        script_id = script[:3]
        script_path = os.path.join(script_directory, script)
        print(f"運行 {script_path}...")

        output = execute_python_script(script_path)
        if output is None:
            continue
        
        try:
            transactions = json.loads(output)
        except json.JSONDecodeError as e:
            print(f"解析腳本 {script} 的輸出時出現錯誤: {e}")
            continue
        
        owner_key = next((key for key in user_info.keys() if key.startswith(script_id)), "未知用戶")
        
        if owner_key not in all_transactions:
            all_transactions[owner_key] = {"transaction": [], "total_amount": 0}
        
        print(f"\n來自 {script} 的交易:")
        print(json.dumps(transactions, indent=4, ensure_ascii=False))

        if isinstance(transactions, dict) and "交易記錄" in transactions and "總金額" in transactions:
            all_transactions[owner_key]["total_amount"] += transactions["總金額"]
            all_transactions[owner_key]["transaction"].extend(transactions["交易記錄"])
        else:
            for transaction in transactions:
                all_transactions[owner_key]["transaction"].append(transaction)
                all_transactions[owner_key]["total_amount"] += transaction.get("金額", 0)

    return all_transactions

def save_and_upload_transaction(owner_key, data, user_info, output_base_directory, sheet, current_month_str):
    """將單個用戶的交易數據保存到Excel並上傳到Google Sheets"""
    output_rows = [
        [
            owner_key,
            transaction.get("時間", "")[:10],
            transaction.get("銀行卡", ""),
            transaction.get("提款方式", ""),
            transaction.get("金額", ""),
            transaction.get("狀態", "")
        ]
        for transaction in data["transaction"]
    ]
    
    df = pd.DataFrame(output_rows, columns=["編號姓名", "時間", "銀行卡", "提款方式", "金額", "狀態"])
    total_row = pd.DataFrame([["", "", "", "總金額", data["total_amount"], ""]], columns=["編號姓名", "時間", "銀行卡", "提款方式", "金額", "狀態"])
    df = pd.concat([df, total_row], ignore_index=True)

    output_file_name = f'{owner_key}_{current_month_str}蝦皮提款記錄.xlsx'
    output_file_path = os.path.join(output_base_directory, output_file_name)

    df.to_excel(output_file_path, index=False)
    format_excel(output_file_path, font_size=20)

    print(f"{owner_key} 的交易已寫入 {output_file_path}")

    market_data = {owner_key[:7]: output_rows}
    upload_to_google_sheets(market_data, sheet, user_info)

def main():
    # 獲取所有腳本
    scripts = [f for f in os.listdir(SCRIPT_DIRECTORY) if f.endswith('requests.py')]
    
    # 加載用戶信息
    user_info_list = load_json_file(USER_INFO_PATH)
    if not user_info_list:
        print("無法加載用戶信息，程序終止。")
        return
    
    # 創建用戶ID到用戶名的映射
    user_info = {f"{user['id']}_{user['owner']}": user['owner'] for user in user_info_list}
    
    # 收集交易數據
    all_transactions = collect_transactions(scripts, SCRIPT_DIRECTORY, user_info)

    # 授權並打開Google Sheets
    gc = pygsheets.authorize(service_file=GOOGLE_SHEETS_KEY)
    sheet = gc.open_by_url(GOOGLE_SHEETS_URL)

    # 獲取當前月份
    utc_now = datetime.utcnow()
    utc_plus_8_now = utc_now + timedelta(hours=8)
    current_month_str = utc_plus_8_now.strftime("%Y-%m")
    output_month_directory = os.path.join(OUTPUT_BASE_DIRECTORY, current_month_str)

    if not os.path.exists(output_month_directory):
        os.makedirs(output_month_directory)

    # 依次處理每個用戶的交易數據
    for owner_key, data in all_transactions.items():
        save_and_upload_transaction(owner_key, data, user_info, output_month_directory, sheet, current_month_str)

    print("所有資料已成功備份到 Google Sheets！")

if __name__ == "__main__":
    main()