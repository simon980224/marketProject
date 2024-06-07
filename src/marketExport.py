import pandas as pd
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# 讀取數據的文件夾路徑
folder_path = '/Users/chenyaoxuan/Desktop/test/'
xlsx_files = glob.glob(folder_path + '*.xlsx')

# 用戶資料
user_data = {
    "d51gtsdkij": {"003": "鐘維嘉"},
    "chi6na8": {"004": "許嘉傑"},
    "qdtfdj4asx": {"022": "蕭喬文"},
    "tookls01": {"027": "王緯德"},
    "csherswa": {"034": "趙偉廷"},
    "0l19pqrp8s": {"042": "劉漢成"},
    "gyk9k3og52": {"043": "陳冠華"},
    "ruili888": {"044": "蕭兆揚"},
    "f0ylv571rl": {"045": "李崇恺"},
    "scxt7aw8ic": {"053": "沈安培"},
    "cr9qpc0org": {"064": "阮羿勲"},
    "irerghmt0p": {"076": "沈政彥"},
    "li5wdw27p1": {"082": "吳祥瑋"},
    "qgo3uvja8e": {"104": "饒傑華"},
    "5v5mkughjk": {"156": "賴慧淑"},
    "bnczq9jjne": {"230": "陳建良"},
    "kp1usft_6x": {"240": "張文彰"},
    "u_gepzjb3x": {"254": "柯英丞"},
    "4ildmehv0p": {"258": "林斯珮"}
}

# 初始化數據字典
processed_data = {}

# 讀取每個 Excel 文件並處理數據
for file in xlsx_files:
    df = pd.read_excel(file)
    result = []
    data_list = df.values.tolist()

    market_acc = data_list[4][1]
    market_name = data_list[5][1]
    market_detail = data_list[17:]

    for detail in market_detail:
        if detail[1] == '已提款金額' and detail[4] == '支出' and detail[6] != '失敗':
            date = detail[0][:10].replace('-', '/')  # 轉換日期格式為 yyyy/mm/dd
            result.append([date, abs(detail[5])])  # 轉換金額為正數

    processed_data[market_acc] = {"name": market_name, "details": sorted(result, key=lambda x: x[0], reverse=False)}

# 按照 user_data 中的 ID 排序
sorted_data_keys = sorted(processed_data.keys(), key=lambda x: list(user_data.get(x, {"999": ""}).keys())[0])

# 初始化 Excel 工作簿
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "未命名的試算表"

# 設置字體、對齊方式
font_black = Font(name='Arial', bold=True, size=12, color="000000")
font_red = Font(name='Arial', bold=True, size=12, color="FF0000")
font_blue = Font(name='Arial', bold=True, size=12, color="0000FF")
header_font = Font(name='Arial', bold=True, size=12, color="FFFFFF")
center_align = Alignment(horizontal="center")

# 設置填充顏色
header_fill_grey = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
fill_grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 灰色填充

start_row = 5
col_offset = 0

# 插入數據到 Excel
for key in sorted_data_keys:
    value = processed_data[key]
    # 獲取編號和名字
    user_info = user_data.get(key, {"未知編號": "未知名字"})
    user_id = list(user_info.keys())[0]
    user_name = list(user_info.values())[0]
    
    # 合併前四格並插入編號和名字，設置為文本格式
    worksheet.merge_cells(start_row=1, start_column=1 + col_offset, end_row=1, end_column=5 + col_offset)
    worksheet.merge_cells(start_row=2, start_column=1 + col_offset, end_row=2, end_column=5 + col_offset)
    
    cell = worksheet.cell(row=1, column=1 + col_offset, value=user_id)
    cell.alignment = center_align
    cell.font = font_black
    worksheet.column_dimensions[get_column_letter(1 + col_offset)].number_format = '@'
    
    cell = worksheet.cell(row=2, column=1 + col_offset, value=user_name)
    cell.alignment = center_align
    cell.font = font_black
    worksheet.column_dimensions[get_column_letter(1 + col_offset)].number_format = '@'
    
    # 插入第3列和第4列的標題
    cell = worksheet.cell(row=3, column=1 + col_offset, value="總計")
    cell.font = header_font
    cell.fill = header_fill_grey
    cell.alignment = center_align
    
    sum_formula = f"=SUM({get_column_letter(2 + col_offset)}{start_row}:{get_column_letter(2 + col_offset)}999)"
    cell = worksheet.cell(row=3, column=2 + col_offset, value=sum_formula)
    cell.font = font_red
    cell.alignment = center_align
    
    cell = worksheet.cell(row=4, column=1 + col_offset, value="日期")
    cell.font = header_font
    cell.fill = header_fill_grey
    cell.alignment = center_align
    
    cell = worksheet.cell(row=4, column=2 + col_offset, value="轉出")
    cell.font = font_red
    cell.alignment = center_align
    
    cell = worksheet.cell(row=3, column=3 + col_offset, value="總計")
    cell.font = header_font
    cell.fill = header_fill_grey
    cell.alignment = center_align
    
    sum_formula = f"=SUM({get_column_letter(4 + col_offset)}{start_row}:{get_column_letter(4 + col_offset)}999)"
    cell = worksheet.cell(row=3, column=4 + col_offset, value=sum_formula)
    cell.font = font_blue
    cell.alignment = center_align
    
    cell = worksheet.cell(row=4, column=3 + col_offset, value="日期")
    cell.font = header_font
    cell.fill = header_fill_grey
    cell.alignment = center_align
    
    cell = worksheet.cell(row=4, column=4 + col_offset, value="轉入")
    cell.font = font_blue
    cell.alignment = center_align
    
    cell = worksheet.cell(row=3, column=5 + col_offset, value="備註")
    cell.font = header_font
    cell.fill = header_fill_grey
    cell.alignment = center_align
    worksheet.merge_cells(start_row=3, start_column=5 + col_offset, end_row=4, end_column=5 + col_offset)
    
    # 插入詳細數據並延伸灰色填充到最底部
    for j in range(995):  # 從第5行到第999行
        row_index = start_row + j
        if j < len(value["details"]):
            entry = value["details"][j]
            date_value = entry[0]
            amount_value = entry[1]
        else:
            date_value = ""
            amount_value = ""
        
        cell = worksheet.cell(row=row_index, column=1 + col_offset, value=date_value)
        cell.font = font_black
        cell.fill = fill_grey
        
        cell = worksheet.cell(row=row_index, column=2 + col_offset, value=amount_value)
        cell.font = font_red
        
        cell = worksheet.cell(row=row_index, column=3 + col_offset, value="")
        cell.font = font_black
        cell.fill = fill_grey
        
        cell = worksheet.cell(row=row_index, column=4 + col_offset, value="")
        cell.font = font_blue
        
        cell = worksheet.cell(row=row_index, column=5 + col_offset, value="")
        cell.font = font_black

    col_offset += 5  # 每個 section 占用 5 列

# 保存工作簿到桌面
output_file_path = '/Users/chenyaoxuan/Desktop/output.xlsx'
workbook.save(output_file_path)